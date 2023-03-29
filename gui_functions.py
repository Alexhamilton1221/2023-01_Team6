import os
import tkinter as tk
from tkinter import *
from tkinter import ttk
import customtkinter
from tkinter import ttk, filedialog
from tkinter import messagebox
import datetime
import openpyxl
from Database.classrooms import Classrooms
from Database.classroom import Classroom
from Database.students import Students
from Database.student import Student
import random
from Database.cohorts import Cohorts
from Database.programs import Programs
from hardCodedClassrooms import temp_Classroom_add
from hardCodedCourses import temp_create_courses
import date_create
import datetime

#from datetime import datetime, timedelta

# Global Vars for preventing multiple windows opening
new_window_open = False
new_window = None

#Track if previous month was Febuary (short)
#is_short=False

#import main as m




#from datetime import datetime, timedelta

#This is for Calendar Creation
global lbl_x,lbl_y
global reg_numbers 
global date_dict
date_dict = date_create.DateCreate(datetime.datetime.now().year, 'Fall')
date_dict.insert_class_days()
reg_numbers = []
global student_info


student_info = Students()
lbl_x=50; lbl_y=20

def import_excel(file_name,imp_type, spn=None):
   global stud_file,res_file
   try:
       file = filedialog.askopenfile(mode='r', filetypes=[('CSV files', '*.xlsx')])

       if file == None:
           return None


       #Checks flag variable to update correct path
       if imp_type==1:
            stud_file=os.path.abspath(file.name)
            registration, student_list = get_registration(stud_file)

            update_spinners(registration, spn)

      
       elif imp_type==2:
            res_file=os.path.abspath(file.name)
            return get_classrooms(file.name)
        
   except Exception as e:
        messagebox.showwarning("Warning", "Failed to upload file. " + str(e))


# Takes list of classroom objects, spinner vars, and the chosen term
# Calls various subroutines to create schedule, add it to classrooms, and insert students
def form_schedule(classroom_list, vars, var_chosen_term):
    global stud_file,res_file #These are the complete paths to the 2 excel files
    global reg_numbers, student_info, date_dict

    cur_semester = 0
    if var_chosen_term.get() == "Fall":
        date_dict = date_create.DateCreate(datetime.datetime.now().year, 'Fall')
        cur_semester = 1
    elif var_chosen_term.get() == "Winter":
        date_dict = date_create.DateCreate(datetime.datetime.now().year, 'Winter')
        cur_semester = 2
    elif var_chosen_term.get() == "Spring/Summer":
        date_dict = date_create.DateCreate(datetime.datetime.now().year, 'Spring')
        cur_semester = 3

    if cur_semester == 0:
        messagebox.showwarning("Warning", "No semester Selected, select a semester")
        return

    date_dict.insert_class_days()

    # Checks if values in spinners match given list, if spinner is higher use that value
    reg_numbers = match_spinners_to_reg(vars, reg_numbers, student_info)

 
    if reg_numbers == None or reg_numbers == []:
        print("EMPTY REGISTRATION")
        return

    programs = Programs(temp_create_courses())
    classrooms = Classrooms(classroom_list)
    students = reg_numbers

    print(students)

    has_made_schedule = False
    time_mods = []
    cohorts = Cohorts()
    fail_array = None
    while has_made_schedule == False:

        has_made_schedule = True
        classrooms.clear_cohorts()
        cohorts.cohorts = []
        fail_array = cohorts.create_cohorts(classrooms, programs, students, cur_semester, time_mods)
        failed_cohorts = cohorts.create_schedules(cur_semester)
        if failed_cohorts != []:
            for cohort in failed_cohorts:
                for mod in time_mods:
                    if mod.program == cohort.program:
                        mod.modifier += 0.1

                        break
                else:
                    time_mods.append(Cohorts.timeModifer(cohort.program))

            has_made_schedule = False
    if fail_array != None:
        outputString = ""
        for fail_data in fail_array:
            if fail_data[4]:
                outputString += "For Program: " + fail_data[0] + " term " + str(fail_data[1]) + "\n " + str(
                    fail_data[2]) + " classes needed of size minimum size " + str(fail_data[3]) + "\n"
            else:
                outputString += "For Program: " + fail_data[0] + " term: " + str(fail_data[1]) + "\n " + str(
                    fail_data[2]) + " labs needed of size minimum size " + str(fail_data[3]) + "\n"

        messagebox.showwarning("Warning", "Too many students for classrooms: \n" + str(outputString))



    print_schedule(classrooms)
    student_info.convert_student_strs_to_objects(programs)
    student_info.add_to_cohorts(programs, cohorts)

    print("Added")



# If value in spinner is not equal to the registraion list of lists, update or append it
def match_spinners_to_reg(spn_vars, reg_nums, students):
    spn_order = ['PCOM 1', 'PCOM 2', 'PCOM 3', 'BCOM 1', 'BCOM 2', 'BCOM 3', 
                 'PM 1', 'PM 2', 'PM 3', 'BA 1', 'BA 2', 'BA 3', 'GL 1', 'GL 2', 
                 'GL 3', 'FS 1', 'FS 2', 'FS 3', 'DXD 1', 'DXD 2', 'DXD 3', 'BK 1', 'BK 2', 'BK 3']    
    
    new_registraions = []
    if reg_nums == None or reg_nums == []:
        reg_nums = []

    # Isolate the spinner entries into a list of list matching reg_nums
    for i, var in enumerate(spn_vars):
        course, term = spn_order[i].split(' ')
        num_students = int(var.get())
        if int(num_students) > 0:
            print([spn_order[i], num_students])
            new_registraions.append([spn_order[i], num_students])


    # For each value from spinner, check to see if there is a matching 
    for new_reg in new_registraions:
        found = False
        for j, reg in enumerate(reg_nums):
            if reg[0] == new_reg[0]:
                if new_reg[1] > reg_nums[j][1]:
                    reg_nums[j][1] = new_reg[1]
                found = True
                break
        if not found:
            print("APPending", new_reg)
            reg_numbers.append(new_reg)


    return reg_numbers


#Function for downloading scedule, maybe need this
def save_schedule():
    print('Downloading Schedule')
    messagebox.showinfo("Note", "Successfully downloaded the Schedule")

#Form the schedule when optionbox is changed
# def form_schedule_screen(frame_t2_background):
   
#     times =["6:00 am", "6:30 am", "7:00 am", "7:30 am", "8:00 am", "8:30 am",
#              "9:00 am", "9:30 am", "10:00 am", "10:30 am", "11:00 am",
#              "11:30 am",
#              "12:00 pm", "12:30 pm", "1:00 pm", "1:30 pm", "2:00 pm", "2:30 pm",
#              "3:00 pm", "3:30 pm",
#              "4:00 pm", "4:30 pm", "5:00 pm", "5:30 pm", "6:00 pm", "6:30 pm",
#              "7:00 pm"]
#     days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday']
    
#     print()
     # Horizontal lines serparating times
    #for i in days:
    #frame_t2_background.create_line(50, 0, 50, 20, fill="red", width=2)



# Updates spinners on main screen when excel file is imported
# Takes list of registration numbers from get_registration() and a list containing the spinner names and the corresponding vars
def update_spinners(registration, spn):

    # Extract spinner keys and vars
    spn_names = spn[0]
    vars = spn[1]

    for key in registration:
        # extract data from registration numbers
        course = key[0].split(" ")[0]
        term = key[0].split(" ")[1]
        num = key[1]

        # Reconstruct data into format used in spinner creation
        spn_name = "spn_" + course.lower() + '_t' + str(term)
        
        
        #Change the var at the same index as the formatted key in the list of spinners
        vars[spn_names.index(spn_name)].set(str(0))
        vars[spn_names.index(spn_name)].set(str(num))
       



# Parse registration numbers from excel file
# Takes a file name of the excel file
# Returns a list of lists formatted as [["{COURSE} {TERM}" NUM_STUDENTS]]
def get_registration(filename):
    global reg_numbers, student_info
     #Open excel file 
    try:
        sheet = openpyxl.load_workbook(filename).worksheets[0]

    except:
        #Error opening file, return None
        return None
    
    student_list = Students()
    registration = {}

    # Check format for registration file
    if sheet['a1'].value == 'Id':         # STUDENT INFORMATION FORMAT

        
        # Create a student object for every row in excel file
        for row in sheet.iter_rows(min_row=2):
            new_student = Student(id= row[0].value,name=row[1].value, term=row[2].value, 
                                  core=row[3].value, program=row[4].value)
            
            
            student_list.students.append(new_student)

        # Sum core and noncore registration numbers
        for student in student_list.students:
            core_key = str(student.core) + ' ' + str(student.term)
            noncore_key = str(student.program) + ' ' + str(student.term)


            # Increment counter for registration for both core and noncore program
            if core_key in registration.keys():
                registration[core_key] += 1
            else:
                registration[core_key] = 1

            if noncore_key in registration.keys():
                registration[noncore_key] += 1
            else:
                registration[noncore_key] = 1


    else: # REGISTRATION NUMBERS FORMAT

        # For each row in the excel file, skipping the header
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == None:
                continue

            course = row[1].value 
            term = row[0].value 
            num = row[2].value

            if num == 0 or num == None:
                continue

            registration[course + " " + str(term)] = int(num)



    #Turn dictionary into a list of lists
    reg_list = []
    for key in registration:
        reg_list.append([key, registration[key]])

    if len(student_list.students) == 0:
        print("Making objects")
        student_list = create_student_objects(reg_list)
        

    reg_numbers = reg_list
    student_info = student_list
    return (reg_list, student_list)



# Takes a registration list of lists and creates dummy Student objects for them
# Returns a Students() object
def create_student_objects(reg_numbers):

    core_courses = ['PCOM', 'BCOM']
    core_registrations = {}
    noncore_registrations = []
    student_list = Students()

    data = reg_numbers.copy()
    id = -1

    # Seperate core and non core registrations
    for item in data:
        course, term = item[0].split(' ')
        num_students = item[1]

        if course in core_courses:
            if course in core_registrations:
                core_registrations[course].append((term, num_students))
            else:
                core_registrations[course] = [(term, num_students)]

        else:
            for i in range(num_students):
                noncore_registrations.append((course, term))

    # For each core registration, find a matching non core, create the student
    for core_course, core_reg in core_registrations.items():
        for term, num_students in core_reg:
            for i in range(num_students):
                for j, (noncore_course, noncore_term) in enumerate(noncore_registrations):
                    if term == noncore_term:
                        new_student = Student(id=id, name="Fname", core=core_course, program=noncore_course, term=term)
                        student_list.students.append(new_student)
                        del noncore_registrations[j]
                        id -= 1
                        break

    return Students






'''
Takes a file name refrencing program information excel file
Returns a Classrooms object containing a list of Classroom objects for each classroom
'''
def get_classrooms(filename):

    #Init excel work sheet
    ws = openpyxl.load_workbook(filename).worksheets[-1]

    #Object to hold each Classroom
    room_list = Classrooms()

    #For each classroom row, skipping the header
    for row in ws.iter_rows(min_row=2):
        if row[0].value == None:
            continue

        #Sterilize room No. info
        room_no = row[0].value
        #Save capacity as an int
        cap = int(row[1].value)
        #Create Classroom object
        new_classroom = Classroom(room_no, cap, ("lab" in row[0].value.lower()))
        #Add to Classrooms List
        room_list.add_classroom(new_classroom)

    #Return entire list
    
    return room_list.classrooms

#Takes a time and converts it to its regular time 
def conv_time(start_time,end_time):
     # Init display variable
    display_time = start_time
    display_time_end = end_time
    start_pm = False; end_pm = False
    start_pm = False; end_pm = False

    # If time is greater than 12, keep in 12hr format
    if display_time >= 13:
        display_time -= 12
        start_pm = True
    if display_time_end >= 13:
        display_time_end -= 12
        end_pm = True

    # If start or end time is a half hour,
    if display_time.is_integer():
        display_time = f"{int(display_time)}:00"
    else:
        display_time = f"{int(display_time)}:30"

    if start_pm:
        display_time += "pm"
    else:
        display_time += "am"

    if display_time_end.is_integer():
        display_time_end = f"{int(display_time_end)}:00"
    else:
        display_time_end = f"{int(display_time_end)}:30"

    if end_pm:
        display_time_end += "pm"
    else:
        display_time_end += "am"

    return display_time,display_time_end


# Takes list of entries from schedule page, time of class as a float, list of days as an index, length as a float
def create_schedule_block(entries_dict, lecture, name, cohort): 

    program = cohort.program.name

    colors = {"BCOM": '#f4ceb8', 'PCOM': '#c2a2c2', 'BA': '#e9a7b8', 'DXD': '#a7bed3',
              'PM': '#00A5E3', 'FS': '#8dd7bf', 'GLM': '#00cdac', 'BK': '#6c88c4'}
    color = ''
    for key in colors.keys():
        if program in key:
            color = colors[key]
    if len(color) == 0:
        r = lambda: random.randint(0,255)
        color = '#%02X%02X%02X' % (r(),r(),r())


    # Schedule begins at 8 so remove those indexes
    starting_hour = lecture.start_time - 8

    # Indexes in hlaf hour increments so double the starting hour
    starting_index = starting_hour // 0.5

    # Get day in terms of index of week, lectures start at 1, indexs at 0 so -1
    day = (lecture.day % 4)-1
    # If wrapped around, 
    if day == -1:
        day = 3

    length =  lecture.end_time - lecture.start_time

    #print(name, lecture.start_time, lecture.day, length)

    # Init display variable
    display_time = lecture.start_time
    display_time_end = lecture.end_time
    start_pm = False; end_pm = False

    # If time is greater than 12, keep in 12hr format
    if display_time >= 13:
        display_time -= 12
        start_pm = True
    if display_time_end >= 13:
        display_time_end -= 12
        end_pm = True

    # If start or end time is a half hour,  
    if display_time.is_integer():
        display_time = f"{int(display_time)}:00"
    else:
        display_time = f"{int(display_time)}:30"

    if start_pm:
        display_time += "pm"
    else:
        display_time += "am"

    if display_time_end.is_integer():
        display_time_end = f"{int(display_time_end)}:00"
    else:
        display_time_end = f"{int(display_time_end)}:30"

    if end_pm:
        display_time_end += "pm"
    else:
        display_time_end += "am"
        
    return display_time,display_time_end

# Takes list of entries from schedule page, time of class as a float, list of days as an index, length as a float
def create_schedule_block(entries_dict, lecture, name, cohort): 

    program = cohort.program.name

    colors = {"BCOM": '#f4ceb8', 'PCOM': '#c2a2c2', 'BA': '#e9a7b8', 'DXD': '#a7bed3',
              'PM': '#00A5E3', 'FS': '#8dd7bf', 'GLM': '#00cdac', 'BK': '#6c88c4'}
    color = ''
    for key in colors.keys():
        if program in key:
            color = colors[key]
    if len(color) == 0:
        r = lambda: random.randint(0,255)
        color = '#%02X%02X%02X' % (r(),r(),r())


    # Schedule begins at 8 so remove those indexes
    starting_hour = lecture.start_time - 8

    # Indexes in half hour increments so double the starting hour
    starting_index = starting_hour // 0.5

    # Get day in terms of index of week, lectures start at 1, indexs at 0 so -1
    day = (lecture.day % 4)-1
    # If wrapped around, 
    if day == -1:
        day = 3

    length =  lecture.end_time - lecture.start_time
    
    display_time,display_time_end=conv_time(lecture.start_time,lecture.end_time)

        #For each entry in range of Length in half hour increments
    for hour in range(int(length/0.5)):

        # Get entry object at [Day_index, (starting_hour + each hour in length)]
        entry = entries_dict[(day, starting_index+hour)]

        #Set color to course specific color
        entry.config(disabledbackground = color)

        # Tries to place the label for the schedule block in the middle of the block
        if hour == int(length)-1:
            entry.config(state=NORMAL)
            entry.insert(0, cohort.name + ' - ' + name)
            entry.config(state=DISABLED)

        if hour == int(length):
            entry.config(state=NORMAL)
            entry.insert(0,f"{display_time} - {display_time_end}")
            entry.config(state=DISABLED)



        
    
# This function is called whenever a Spinbox is updated to print the new totals.
def update_totals(spinners,total_labels,row_num,spinner_object):

    all_programs=['pcom','bcom','pm','ba','gl','fs','dxd','bk']
    sum=0
    if row_num==1:
        substring = 'pcom'
    elif row_num==2:
        substring = 'bcom'
    elif row_num==3:
        substring = 'pm'
    elif row_num==4:
        substring = 'ba'    
    elif row_num==5:
        substring = 'gl'    
    elif row_num==6:
        substring = 'fs'    
    elif row_num==7:
        substring = 'dxd'    
    elif row_num==8:
        substring = 'bk'    
    
    for spn in spinner_object: #Error Checking
        #print('test',int(spn.get()))
        if int(spn.get())>100:
            messagebox.showerror("Error", "You entered too many Students. \nA maximum of 100 students is permitted.")
            spn.set(0)
            indices = [i for i, s in enumerate(spinners) if substring in s]
            total_labels[row_num-1].configure(text=0)
            for i in indices:
                text=int(spinner_object[i].get())+total_labels[row_num-1].cget("text")
                #print(type(sum))
                sum+=text
                #print(sum)
            total_labels[row_num-1].configure(text=sum)            
            return
    
    indices = [i for i, s in enumerate(spinners) if substring in s]
      
    total_labels[row_num-1].configure(text=0)

    for i in indices:
        text=int(spinner_object[i].get())+total_labels[row_num-1].cget("text")
            
        total_labels[row_num-1].configure(text=text)

# This function is called when somebody types in a value. Need to check all spinboxes.
def update_all_totals(spn_core,spn_noncore,total_labels,spinner_object):
    #programs=['pcom','bcom','pm','ba','gl','fs','dxd','bk']
    core_programs=['pcom','bcom']; non_core_programs=['pm','ba','gl','fs','dxd','bk']
    row_num=1
    
    
    
    for core_substring in core_programs:
        
        core_indices = [i for i, s in enumerate(spn_core) if core_substring in s]
        
        #print(core_indices)

        total_labels[row_num-1].configure(text=0)
        for i in core_indices:
            text=int(spinner_object[i].get())+total_labels[row_num-1].cget("text")
                
            total_labels[row_num-1].configure(text=text)
        #print(row_num)
        row_num+=1



        
    #print("break")

    #row_num=1
    for non_core_substring in non_core_programs:
        #print(non_core_substring)

        non_core_indices = [j for j, k in enumerate(spn_noncore) if non_core_substring in k]
        
        for i in range(len(non_core_indices)): #Deal with new list by adding 6
            non_core_indices[i] += 6
        
        #print(non_core_indices)

        total_labels[row_num-1].configure(text=0)

        for j in non_core_indices:
            text=int(spinner_object[j].get())+total_labels[row_num-1].cget("text")
                
            total_labels[row_num-1].configure(text=text)
        
        #print(row_num)
        row_num+=1
    
    # full_spn=spn_core.update(spn_noncore)
    # print('HERE',type(spn_core))
    # print(spn_core)
    # print(spn_noncore)

    # print('HERE',type(full_spn))
    # for substring in programs:
    
    #     indices = [i for i, s in enumerate(full_spn) if substring in s]
      
    #     total_labels[row_num-1].configure(text=0)

    #     for i in indices:
    #         text=int(spinner_object[i].get())+total_labels[row_num-1].cget("text")
            
    #         total_labels[row_num-1].configure(text=text)

        
def change_classroom(label, var):
    data = var.get()
    label.configure(text=str(data))
   
   
   
# #Get current Season for Term
# def get_season():
#     now = datetime.datetime.now()
#     if now.month >= 9 and now.month <= 12:
#         return "Fall"
#     elif now.month >= 6 and now.month <= 8:
    
#     else:
#         return "Winter"   
#     # if now.month >= 3 and now.month <= 5:
#     #     return "Spring"
#     # elif now.month >= 6 and now.month <= 8:
#     #     return "Summer"
#     # elif now.month >= 9 and now.month <= 11:
#     #     return "Fall"
#     # else:
#     #     return "Winter"   
# def on_enter(event):
#     event.widget.config(bg="#3e3e42")

# def on_leave(event):
#     event.widget.config(bg="#252526")


def clear_schedule(entries):
    for index in entries:
        entries[index].config(state=NORMAL)
        entries[index].delete(0, END)
        entries[index].config(disabledbackground = '#ffffff')
        entries[index].config(state=DISABLED)

#prints total schedule for a room
def print_schedule(classrooms):
    #for room in classrooms:
    for room in classrooms.classrooms:
        for cohort in room.cohorts:
            for course in cohort.courses:
                for lecture in course.lectures:
                    if lecture.day < 12:
                        print(room.name, ' - ', course.name, lecture.day, lecture.start_time, course.delivery)
                        
                        
def print_cohorts(classrooms,cohort_name,text_field):

    #Testing Print
    text_field.configure(state='normal')

    text_field.delete("1.0", "end") #Clear Text Field
    for i,x in enumerate(classrooms.classrooms):
        if classrooms.classrooms[i].name == cohort_name:
            for cohort in classrooms.classrooms[i].cohorts:
                text_field.insert(tk.END, cohort.name + "\nClassroom: " + str(cohort.room) + " Lab: " + str(cohort.lab) + "\n")
                text_field.insert(tk.END, "Students: \n")
                count = 0
                student_info = ""
                for student in cohort.students:
                    student_info += "    Id: " + str(student.id) + "    name: " + student.name + ", "
                    count += 1
                    if count == 5:
                        text_field.insert(tk.END, student_info + '\n')
                        student_info = ""
                        count = 0

                text_field.insert(tk.END, "Courses: \n")
                for course in cohort.courses:

                    lecture = course.lectures[0]
                    end_lecture = course.lectures[len(course.lectures) - 1]
                    days_spacing = " - "
                    # if lecture.day<10:
                    #     days_spacing="   -"
                    # elif lecture.day<99:
                    #     days_spacing=" -"
                    # else:
                    #     days_spacing="-"

                    #Init display variable
                    display_time = lecture.start_time
                    display_time_end = lecture.end_time
                    start_pm = False; end_pm = False

                        # If time is greater than 12, keep in 12hr format
                    if display_time > 12:
                        display_time -= 12
                        start_pm = True
                    if display_time_end > 12:
                        display_time_end -= 12
                        end_pm = True

                      # If start or end time is a half hour,
                    if display_time.is_integer():
                        display_time = f"{int(display_time)}:00"
                    else:
                        display_time = f"{int(display_time)}:30"

                    if start_pm:
                        display_time += "pm"
                    else:
                        display_time += "am"

                    if display_time_end.is_integer():
                        display_time_end = f"{int(display_time_end)}:00"
                    else:
                        display_time_end = f"{int(display_time_end)}:30"

                    if end_pm:
                        display_time_end += "pm"
                    else:
                        display_time_end += "am"


                    if len(display_time)==6:
                        display_time=f"0{display_time} "

                        #For testing include classroom name but remove later
                    text_field.insert(tk.END,course.name+' - Days: '+str(lecture.day)+str(days_spacing)+str(end_lecture.day)+
                    '      Time: '+str(display_time) + " - " + str(display_time_end)+'      Delivery Type: '+ course.delivery+'\n')
                text_field.insert(tk.END, '   \n')


    text_field.configure(state='disabled')
        
class Calendar(tk.Frame):
    def __init__(self, parent,array_rect,array_lbl,semester_lectures):
        tk.Frame.__init__(self, parent)
        # Create Array for all Rectangles
        # When rectangles are saved to array, they are saved by an ID number
        # Must reference them using methods.
        self.array_rect=array_rect
        self.array_lbl=array_lbl
        self.semester_lectures=semester_lectures
        # Create a vertical scrollbar
        scrollbar = ttk.Scrollbar(self, orient='vertical')
        scrollbar.pack(side='right', fill='y')

        # Create a canvas to contain the widgets
        canvas = tk.Canvas(self, bd=0, highlightthickness=0, yscrollcommand=scrollbar.set)
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=canvas.yview)

        # Set the canvas to expand to fill the entire frame
        self.canvas = canvas
        canvas.bind('<Configure>', self._configure_canvas)

        # Create a frame to hold the widgets
        self.inner_frame = tk.Frame(canvas)
        self.inner_frame_id = canvas.create_window((0, 0), window=self.inner_frame, anchor='nw')

        # Hide the canvas
        canvas.configure(borderwidth=0, highlightthickness=0)
            
        
    def _configure_canvas(self, event):
        self.canvas.config(scrollregion=self.canvas.bbox('all'))

    def update_viewport(self):
        self.canvas.config(scrollregion=self.canvas.bbox('all'))
    
    def close_new_window(self):
        global new_window_open
        # Set the new window flag to False and destroy the window
        new_window_open = False
        new_window.destroy()
        
    
    def clean_array(self):
        while self.semester_lectures:
            self.semester_lectures.pop()

    #def clean_array(self):
     #   self.semester_lectures = [subarr for subarr in self.semester_lectures if subarr]
    
    
    def calendar_entry_clicked(self,event,row,col):
        global new_window_open, new_window
        
        new_window_body=("Arial", 14) 

        #Number of columns
        num_of_columns=4
        index=  row *num_of_columns  + col

        # Create New Window Displaying Class Info
        # TODO Make the title the appropriate date
        if not new_window_open:
            #Create New Window
            new_window = tk.Toplevel(self.canvas)
            new_window.title("Date") 
            new_window.geometry("640x360")
            new_window.config(background='#252526')
            
            #Call Class to construct Canvas
            entry_frame = Day(new_window)
            entry_frame.place(relx=0.5, rely=0.1, relwidth=1, relheight=0.9, anchor='n')    
            
            indexed_lecs=[]
            for subarr in self.semester_lectures:
                if subarr[0]==index:
                    indexed_lecs.append(subarr)
                    #print('new',subarr)

            sorted_list = sorted(indexed_lecs, key=lambda x: x[5])
            
            #For Testing
            #TODO Place Date in this Object 
            #This is just for testing, replace this later
            try:
                title = tk.Label(new_window, text=f"Date {sorted_list[0][0]}", bg="#252526", fg='white',font=new_window_body)
                title.pack()
            except:
                title = tk.Label(new_window, text=f"Date", bg="#252526", fg='white',font=new_window_body)
                title.pack()


            
            x1=325 ; y1=20 ; pad_y=0
            for class_etr in sorted_list:
                new_text=f" {class_etr[1]} {class_etr[2]} {class_etr[3]} - {class_etr[4]}"
                
                entry_frame.canvas.create_text(x1,y1+pad_y,text=new_text,fill="white",font=new_window_body)
                #self.canvas.itemconfigure(text, fill="blue")

                pad_y+=20
                  

            #Screen Settings
            
            # Make it so that window cannot change size/shape
            new_window.attributes('-fullscreen', False)
            new_window.resizable(False, False)

            #Call function to close window to prevent multiple open windows
            new_window_open = True
            new_window.protocol("WM_DELETE_WINDOW", self.close_new_window)

    
    
    def setup_grid(self):
        rect_size = 180
        count=1
        for row in range(5):
            for col in range(4):
                x1 = col * rect_size
                y1 = row * rect_size
                x2 = x1 + rect_size
                y2 = y1 + rect_size
                rect=self.canvas.create_rectangle(x1, y1, x2, y2, outline='black', fill='white')
                count+=1

                self.canvas.tag_bind(count, "<Button-1>", lambda event, row=row, col=col: self.calendar_entry_clicked(event,row, col+1))

                #self.canvas.tag_bind(rect, "<Button-1>", lambda event, index=count: self.on_rectangle_click())
                #self.canvas.tag_bind(rect, "<Button-1>", self.calendar_entry_clicked)
                self.array_rect.append(rect)

        #Deal With last Entry
        #self.canvas.tag_bind(count+1, "<Button-1>", lambda event, row=row, col=col: self.calendar_entry_clicked(event,row, col+1))

        #print(self.array_rect)
        
    def clear_grid(self):
        text_items = self.canvas.find_all()     
        for item in text_items:
            if self.canvas.type(item) == "text":
                self.canvas.delete(item)


    def get_date(self,date,number):
        new_date = date + datetime.timedelta(days=number-1)
        date_str = new_date.strftime('%B %d, %Y')
        return date_str
    
    
    def calendar_day_entry(self,sorted_list,day_in_month,month_end,prev_month_lengths,start_date_obj,month_start,current_mon_length):
        count=0; rect_size = 180
        #global is_short
    
        # Font for Calendar Creation
        my_font = ("Arial", 24) 
                
        count += prev_month_lengths; date_diff=0
        extra_days=0
        #Accounting for change in months since loop resets
        print('Current Month is:',current_mon_length)
        

        
        #if it isnt first month of term or febuary
        if (month_end>1):
            extra_days=((month_end-1)*5 + 7*(month_end-1))
        
        
        #print(sorted_list)
        for row in range(5):
            if row>=1:
                #This accounts for weekends
                date_diff=3*row
            for col in range(4):
                #print(f'TESTING Count: {count} day_in_month {day_in_month}')             
                count+=1
                pad_y=0
               
                # Default Case: Print all Lectures
                if count==day_in_month and len(sorted_list)<8:
                    #print(f"The Current Day is {day_in_month} and count is {count}")

                    # Add Date Title to Grid Tiles
                    x1 = col * rect_size; x1+=85
                    y1 = row * rect_size+pad_y; y1+=15; 
                    
                    date_str=self.get_date(start_date_obj,count+date_diff+extra_days)
                    
                    new_text=f"{date_str}"

                    text = self.canvas.create_text(x1,y1+pad_y,text=new_text, font=("Helvetica", 10, "bold"))
                        
                    self.array_lbl.append(text)
                    pad_y+=25
                
                    for cal_etr in sorted_list:

                        new_text=f"{cal_etr[0]} {cal_etr[1]} {cal_etr[2]}"

                        text = self.canvas.create_text(x1,y1+pad_y,text=new_text)
                        
                        self.array_lbl.append(text)
                        pad_y+=15
                
                # Second Case: Print 8 lectures and add ...
                elif count==day_in_month: 
                    x1 = col * rect_size; x1+=85
                    y1 = row * rect_size+pad_y; y1+=15; 
                    
                    # Add Date Title to Grid Tiles
                    x1 = col * rect_size; x1+=85
                    y1 = row * rect_size+pad_y; y1+=15; 
                    
                    date_str=self.get_date(start_date_obj,count+date_diff-(month_start-1))
                    
                    new_text=f"{date_str}"

                    text = self.canvas.create_text(x1,y1+pad_y,text=new_text, font=("Helvetica", 10, "bold"))
                        
                    self.array_lbl.append(text)
                    pad_y+=25
                    

                    for cal_etr in sorted_list[:8]:
                        new_text=f"{cal_etr[0]} {cal_etr[1]} {cal_etr[2]}"

                        text = self.canvas.create_text(x1,y1+pad_y,text=new_text)
                        self.array_lbl.append(text)
                        pad_y+=15
                    text = self.canvas.create_text(x1,y1+pad_y,text="...", font=my_font)
        is_short=True
        
       
    def weekend_entry(self,i):
        count=0; rect_size = 180

        
        for row in range(5):
            for col in range(7):
                count+=1
                pad_y=0
                
                # Default Case: Print all Lectures
                if count==i:
                    x1 = col * rect_size; x1+=85
                    y1 = row * rect_size+pad_y; y1+=15; 

                    
                    new_text=f"WEEKEND"

                    text = self.canvas.create_text(x1,y1+pad_y,text=new_text)
                        
                    self.array_lbl.append(text)
                    pad_y+=15
                
             
def reset_is_short():
    global is_short
    is_short=False
    print('reseting')





class Day(tk.Frame):
    def __init__(self, parent,):
        tk.Frame.__init__(self, parent)
        
        # Create a vertical scrollbar
        scrollbar = ttk.Scrollbar(self, orient='vertical')
        scrollbar.pack(side='right', fill='y')

        # Create a canvas to contain the widgets
        canvas = tk.Canvas(self, bd=0, highlightthickness=0, yscrollcommand=scrollbar.set,background='#3e3e42')
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=canvas.yview)

        # Set the canvas to expand to fill the entire frame
        self.canvas = canvas
        canvas.bind('<Configure>', self._configure_canvas)

        # Create a frame to hold the widgets
        self.inner_frame = tk.Frame(canvas)
        self.inner_frame_id = canvas.create_window((0, 0), window=self.inner_frame, anchor='nw')

        # Hide the canvas
        canvas.configure(borderwidth=0, highlightthickness=0)
            
        
    def _configure_canvas(self, event):
        self.canvas.config(scrollregion=self.canvas.bbox('all'))

    def update_viewport(self):
        self.canvas.config(scrollregion=self.canvas.bbox('all'))

        # lbl_x+=100
        #lbl_y+=20



def reset(classroom_list, spn_vars, spn_core,spn_noncore,total_labels,spinner_object):
    global reg_numbers
    reg_numbers = []
    for room in classroom_list:
        room.cohorts = []

    for var in spn_vars:
        var.set(0)

    update_all_totals(spn_core,spn_noncore,total_labels,spinner_object)
    return []


def term_stats(var_chosenterm,var_dispmonth_calendar):
    #Can add leap year later if needed
    # days_in_month = {'January': 31,'February': 28, 'March': 31,'April': 30,'May': 31,
    # 'June': 30,'July': 31,'August': 31,'September': 30,'October': 31,'November': 30,'December': 31 }
    
    # fall_months = {'September': 30,'October': 31,'November': 30,'December': 31 }
    # winter_months = {'January': 31,'February': 28, 'March': 31,'April': 30}
    # springsum_months = {'May': 31,'June': 30,'July': 31}
    
    #Rough estimate, subtract fri, sat, sun, doest account for holidays
    fall_months = {'September':18,'October': 19,'November': 18,'December': 19 }
    winter_months = {'January': 19,'February': 16, 'March': 19,'April': 18}
    springsum_months = {'May': 19,'June': 18,'July': 19,'August':19}
    
    start_date_obj=''
    month_start=1; month_end=0; current_mon=1 ;prev_month_lengths=0;current_mon_length=0
    #Get Current month to find which term
    current_term=var_chosenterm.get()
    current_month=var_dispmonth_calendar.get()

    print('The current term is ',current_term)
    print('The current month is ',current_month)
    
    if current_term=="Fall":
        print("Fall Time")
        start_date_obj = datetime.date(2023, 9, 1)
        
        for key,value in fall_months.items():
            if key != current_month: #If it is not the right month
                month_start+=value
                month_end+=value
                current_mon+=1
                prev_month_lengths+=value
            else:
                month_end+=value
                #print(f"Current Month {key} Month Start: {month_start+1} Days {value} Total Days Elapsed {day_gap}")

                #Return legnth of current month
                current_mon_length=value
                return month_start,month_end,current_mon,prev_month_lengths,start_date_obj,current_mon_length
            
    elif current_term=="Winter":
        print("Winter Time")
        start_date_obj = datetime.date(2024, 1, 1)

        for key,value in winter_months.items():
            if key != current_month: #If it is not the right month
                month_start+=value
                month_end+=value
                current_mon+=1
                prev_month_lengths+=value

            else:
                month_end+=value
                #print(f"Current Month {key} Month Start: {month_start+1} Days {value} Total Days Elapsed {day_gap}"
                #Return legnth of current month
                current_mon_length=value
                return month_start,month_end,current_mon,prev_month_lengths,start_date_obj,current_mon_length
            
    elif current_term=="Spring/Summer":
        print("Spring/Summer Time")
        start_date_obj = datetime.date(2024, 5, 1)

        for key,value in springsum_months.items():
            if key != current_month: #If it is not the right month
                month_start+=value
                month_end+=value
                current_mon+=1
                prev_month_lengths+=value

            else:
                month_end+=value
                print(f"Current Month {key} Month Start: {month_start+1} Days {value}")
                #Remove first montrh
                #Return legnth of current month
                current_mon_length=value
                
                return month_start,month_end,current_mon,prev_month_lengths,start_date_obj,current_mon_length
    return month_start,month_end,current_mon,prev_month_lengths,start_date_obj,current_mon_length



def update_schedule_labels(labels, week):
    global date_dict


    week_start_day = (((week-1)*4)+1)-2



    for i, label in enumerate(labels):
        if week_start_day+i == 0 or week_start_day+i == -1:
            day_num = date_dict.locate_start_day()+week_start_day+i
            txt = f"{month_num_to_name(str(list(date_dict.calendar_dictionary.keys())[0]))}"
            txt += f" {date_suffix(day_num)}"


        else:
            txt = schedule_day_to_date(date_dict, week_start_day+i)

        label.configure(text=txt)

def date_suffix(date):
    suffix_dict = {1: "1st",
               2: "2nd",
               3: "3rd",
               4: "4th",
               5: "5th",
               6: "6th",
               7: "7th",
               8: "8th",
               9: "9th",
               10: "10th",
               11: "11th",
               12: "12th"}

    return suffix_dict[date]

def month_num_to_name(num):
    months = {	'1':'Janauary',
		'2':'February',
		'3':'March',
		'4':'April',
		'5':'May',
		'6':'June',
		'7':'July',
		'8':'August',
		'9':'September',
		'10':'October',
		'11':'November',
		'12':'December'		}

    return months[num]



def schedule_day_to_date(date_dict, day):




    if day in [0,-1]:
        txt = month_num_to_name(date_dict.calendar_dictionary.keys()[0])
        txt += f"{(date_dict.locate_start_day())+day}"




    for month in date_dict.calendar_dictionary:

        for date in date_dict.calendar_dictionary[month]:
            if date_dict.calendar_dictionary[month][date] == day:
                date_num = int(date.split('-')[-1])
                text =  f"{month_num_to_name(str(month))} {date_num}"

                if date_num == 1:
                    text += "st"
                elif date_num == 2:
                    text += "nd"
                elif date_num == 3:
                    text += "rd"
                else:
                    text += "th"

                return text



def value_to_key(dict, val):

    for key in dict:
        if dict[key] == val:
            return key
    print("Not Found", val)
    return False


