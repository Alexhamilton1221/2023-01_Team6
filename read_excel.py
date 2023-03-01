import openpyxl
from Database.program import Program
from Database.course import Course
from Database.classroom import Classroom
from Database.classrooms import Classrooms
from Database.cohort import Cohort
from Database.cohorts import Cohorts

import string

'''
Takes a file name refrencing program information excel file
Returns a Classrooms object containing a list of Classroom objects for each classroom
'''
def get_classrooms(filename):

    #Init excel work sheet
    ws = openpyxl.load_workbook(filename).worksheets[4]

    #Object to hold each Classroom
    room_list = Classrooms()

    #For each classroom row, skipping the header
    for row in ws.iter_rows(min_row=2):
        #Sterilize room No. info
        room_no = row[0].value.split(' ')[0]
        #Save capacity as an int
        cap = int(row[1].value)
        #Create Classroom object
        new_classroom = Classroom(room_no, cap, ("lab" in row[0].value.lower()))
        #Add to Classrooms List
        room_list.add_classroom(new_classroom)

    #Return entire list
    return room_list



def get_registration(filename):

     #Open excel file and init return object
    try:
        sheet = openpyxl.load_workbook(filename).worksheets[0]

    except:
        #Error opening file, return None
        return None
    
    registration = {}

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == None:
            continue

        #Idex of values unsure - template not uploaded yet
        course = row[0].value 
        term = row[1].value 
        num = row[2].value

        registration[course + " " + str(term)] = int(num)


    return registration






#Returns the index of a letter in the alphabet
def char_position(letter):
    return ord(letter.lower()) - 97
    
